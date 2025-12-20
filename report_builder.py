import io
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Dict, List, Tuple, Union, Iterable

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from csv_loader import REPORT_TYPE_ROMAN, load_source_dataframe_for_report_type
from report_schema import get_source_columns_map, load_config, sanitize_for_filename


_CLIENT_NUMERIC_2DP_COLS = {
    "Fatturato_lordo_sconto_cassa_CY",
    "Fatturato_lordo_sconto_cassa_PY",
    "Delta_CYvsPY",
    "DeltaPerc_CYvsPY",
    "ORDINATO_INEVASO_CY",
}

_NUM_FORMAT_2DP = "#,##0.00"
_COL_WIDTH_NUM = 18
_TOTAL_LABEL = "Totale"


def parse_report_type_key(report_type: Union[int, str]) -> Tuple[int, str]:
    if isinstance(report_type, int):
        return report_type, str(report_type)

    s = str(report_type).strip()
    if not s:
        raise KeyError("Empty report_type")

    if "_" in s:
        base_str, _ = s.split("_", 1)
        base = int(base_str)
        return base, s

    base = int(s)
    return base, s


def get_logical_fields_specific_first(config: Dict[str, Any], report_type: Union[int, str]) -> List[str]:
    report_types = config["report_types"]
    _, type_key = parse_report_type_key(report_type)

    if type_key not in report_types:
        raise KeyError(f"Unknown report_type: {type_key}")

    specific = report_types[type_key]["specific_fields"]
    shared = config["shared_fields"]

    result: List[str] = []
    seen = set()

    for field in list(specific) + list(shared):
        if field not in seen:
            seen.add(field)
            result.append(field)

    return result


def ensure_unique_columns(df: pd.DataFrame) -> None:
    if not df.columns.is_unique:
        dupes = df.columns[df.columns.duplicated()].astype(str).tolist()
        raise ValueError(f"Duplicate column names detected: {dupes}")


def get_col_idx_1based(df: pd.DataFrame, column_name: str) -> int:
    loc = df.columns.get_loc(column_name)
    if not isinstance(loc, int):
        raise ValueError(
            f"Expected unique column match for '{column_name}', got {type(loc).__name__}. "
            "This usually indicates duplicate column names."
        )
    return loc + 1


def merge_vertical_column(
    ws,
    df: pd.DataFrame,
    column_name: str,
    *,
    exclude_last_row: bool = False,
) -> None:
    if column_name not in df.columns:
        return
    if len(df) <= 1:
        return

    ensure_unique_columns(df)

    col_idx = get_col_idx_1based(df, column_name)
    start_row = 2

    data_len = len(df) - (1 if exclude_last_row else 0)
    if data_len <= 1:
        return

    end_row = start_row + data_len - 1

    ws.merge_cells(
        start_row=start_row,
        start_column=col_idx,
        end_row=end_row,
        end_column=col_idx,
    )

    cell = ws.cell(row=start_row, column=col_idx)
    cell.alignment = Alignment(vertical="top")


def merge_vertical_columns(
    ws,
    df: pd.DataFrame,
    column_names: List[str],
    *,
    exclude_last_row: bool = False,
) -> None:
    for col in column_names:
        merge_vertical_column(ws, df, col, exclude_last_row=exclude_last_row)


def parse_decimal_2dp(value: Any) -> Decimal | None:
    if value is None:
        return None
    if pd.isna(value):
        return None

    s = str(value).strip()
    if not s:
        return None

    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return None

    return d.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def coerce_numeric_2dp_columns(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if col in _CLIENT_NUMERIC_2DP_COLS:
            df[col] = df[col].map(parse_decimal_2dp)
    return df


def _iter_decimals(values: Iterable[Any]) -> Iterable[Decimal]:
    for v in values:
        if v is None or pd.isna(v):
            continue
        if isinstance(v, Decimal):
            d = v
        else:
            d = parse_decimal_2dp(v)
        if d is not None:
            yield d


def _decimal_sum(values: Iterable[Any]) -> Decimal | None:
    total = Decimal("0.00")
    has_any = False
    for d in _iter_decimals(values):
        total += d
        has_any = True
    if not has_any:
        return None
    return total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _decimal_mean(values: Iterable[Any]) -> Decimal | None:
    total = Decimal("0.00")
    count = 0
    for d in _iter_decimals(values):
        total += d
        count += 1
    if count == 0:
        return None
    return (total / Decimal(count)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def aggregation_grouping_columns_for_report_type(config: Dict[str, Any], report_type: Union[int, str]) -> List[str]:
    """
    Defines the target dataframe grain for each report type.
    This step eliminates duplicates at export-level by aggregating upstream finer-grain rows.
    """
    fields = config["fields"]

    def sc(logical_name: str) -> str:
        return fields[logical_name]["source_column"]

    _, type_key = parse_report_type_key(report_type)

    if type_key == "1":
        return [sc("ragione_sociale"), sc("gruppo_merceologico")]
    if type_key == "2_CONS":
        return [sc("consorzio"), sc("gruppo_merceologico")]
    if type_key == "2_GRUPPO":
        return [sc("gruppo_commerciale"), sc("gruppo_merceologico")]
    if type_key == "3_CONS":
        return [sc("consorzio"), sc("ragione_sociale"), sc("gruppo_merceologico")]
    if type_key == "3_GRUPPO":
        return [sc("gruppo_commerciale"), sc("ragione_sociale"), sc("gruppo_merceologico")]
    if type_key == "4":
        return [sc("agenzia_anagrafica"), sc("ragione_sociale")]
    if type_key == "5":
        return [sc("agenzia_anagrafica"), sc("ragione_sociale"), sc("gruppo_merceologico")]

    return []


def aggregate_to_report_grain(df: pd.DataFrame, *, group_cols: List[str]) -> pd.DataFrame:
    if df.empty or not group_cols:
        return df

    missing = [c for c in group_cols if c not in df.columns]
    if missing:
        raise RuntimeError(f"Aggregation requires grouping columns missing from dataframe: {missing}")

    # Guard against silently dropping/choosing values for non-numeric columns outside the grouping keys.
    extra_non_numeric = [c for c in df.columns if c not in group_cols and c not in _CLIENT_NUMERIC_2DP_COLS]
    if extra_non_numeric:
        raise RuntimeError(
            "Aggregation would drop/merge non-numeric columns not included in grouping keys. "
            f"Add them to group_cols or define an aggregation strategy. Columns: {extra_non_numeric}"
        )

    agg: Dict[str, Any] = {}
    for col in df.columns:
        if col in group_cols:
            continue
        if col not in _CLIENT_NUMERIC_2DP_COLS:
            continue

        if col == "DeltaPerc_CYvsPY":
            agg[col] = _decimal_mean
        else:
            agg[col] = _decimal_sum

    if not agg:
        return df.drop_duplicates(subset=group_cols, keep="first").reset_index(drop=True)

    df_out = (
        df.groupby(group_cols, dropna=False, sort=False, as_index=False)
        .agg(agg)
        .reset_index(drop=True)
    )

    return df_out


def build_totals_row(
    df: pd.DataFrame,
    *,
    label_col: str,
    keep_cols: List[str],
) -> Dict[str, Any]:
    totals: Dict[str, Any] = {c: None for c in df.columns}

    if not df.empty:
        first = df.iloc[0]
        for c in keep_cols:
            if c in df.columns:
                totals[c] = first[c]

    if label_col in df.columns:
        totals[label_col] = _TOTAL_LABEL
    else:
        fallback = next((c for c in df.columns if c not in _CLIENT_NUMERIC_2DP_COLS), df.columns[0])
        totals[fallback] = _TOTAL_LABEL

    for col in df.columns:
        if col not in _CLIENT_NUMERIC_2DP_COLS:
            continue

        if col == "DeltaPerc_CYvsPY":
            total = Decimal("0.00")
            count = 0

            for v in df[col]:
                if v is None or pd.isna(v):
                    continue

                if isinstance(v, Decimal):
                    parsed = v
                else:
                    parsed = parse_decimal_2dp(v)

                if parsed is None:
                    continue

                total += parsed
                count += 1

            if count == 0:
                totals[col] = None
            else:
                mean = (total / Decimal(count)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                totals[col] = mean

            continue

        total = Decimal("0.00")
        for v in df[col]:
            if v is None or pd.isna(v):
                continue
            if isinstance(v, Decimal):
                total += v
            else:
                parsed = parse_decimal_2dp(v)
                if parsed is not None:
                    total += parsed

        totals[col] = total.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    return totals


def append_totals_row(
    df: pd.DataFrame,
    *,
    label_col: str,
    keep_cols: List[str],
) -> pd.DataFrame:
    if df.empty:
        return df

    totals = build_totals_row(df, label_col=label_col, keep_cols=keep_cols)
    return pd.concat([df, pd.DataFrame([totals])], ignore_index=True)


def insert_subtotals(
    df: pd.DataFrame,
    *,
    subtotal_cols: List[str],
    label_col: str,
) -> pd.DataFrame:
    if df.empty or not subtotal_cols:
        return df

    for c in subtotal_cols:
        if c not in df.columns:
            return df

    out_frames: List[pd.DataFrame] = []
    keep_cols = subtotal_cols[:-1]  # keep the parent keys

    df_base = df.reset_index(drop=True)
    for _, df_sub in df_base.groupby(subtotal_cols, dropna=False, sort=False):
        out_frames.append(df_sub)
        out_frames.append(pd.DataFrame([build_totals_row(df_sub, label_col=label_col, keep_cols=keep_cols)]))

    return pd.concat(out_frames, ignore_index=True)


def apply_client_numeric_formatting(ws, df: pd.DataFrame) -> None:
    target_cols = [c for c in df.columns if c in _CLIENT_NUMERIC_2DP_COLS]
    if not target_cols or df.empty:
        return

    ensure_unique_columns(df)

    start_row = 2
    end_row = start_row + len(df) - 1

    for col_name in target_cols:
        col_idx_1based = get_col_idx_1based(df, col_name)

        for r in range(start_row, end_row + 1):
            cell = ws.cell(row=r, column=col_idx_1based)
            if cell.value is None or cell.value == "":
                continue
            cell.number_format = _NUM_FORMAT_2DP

        col_letter = get_column_letter(col_idx_1based)
        current_width = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = max(current_width or 0, _COL_WIDTH_NUM)


def apply_totals_row_bold(ws, df: pd.DataFrame) -> None:
    if df.empty:
        return

    max_row = 1 + len(df)  # header row is 1
    max_col = len(df.columns)
    bold_font = Font(bold=True)

    for r in range(2, max_row + 1):
        is_total = False
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value == _TOTAL_LABEL:
                is_total = True
                break

        if not is_total:
            continue

        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).font = bold_font


def to_excel_bytes(
    df: pd.DataFrame,
    sheet_name: str,
    merge_columns: List[str] | None = None,
) -> bytes:
    ensure_unique_columns(df)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        if merge_columns:
            merge_vertical_columns(ws, df, merge_columns, exclude_last_row=False)

        apply_client_numeric_formatting(ws, df)
        apply_totals_row_bold(ws, df)

    buf.seek(0)
    return buf.read()


def grouping_columns_for_report_type(config: Dict[str, Any], report_type: Union[int, str]) -> List[str]:
    fields = config["fields"]

    def sc(logical_name: str) -> str:
        return fields[logical_name]["source_column"]

    _, type_key = parse_report_type_key(report_type)

    if type_key == "1":
        return [sc("ragione_sociale")]
    if type_key == "2_CONS":
        return [sc("consorzio")]
    if type_key == "2_GRUPPO":
        return [sc("gruppo_commerciale")]
    if type_key == "3_CONS":
        return [sc("consorzio"), sc("ragione_sociale")]
    if type_key == "3_GRUPPO":
        return [sc("gruppo_commerciale"), sc("ragione_sociale")]
    if type_key == "4":
        return [sc("agenzia_anagrafica")]
    if type_key == "5":
        return [sc("agenzia_anagrafica"), sc("ragione_sociale")]
    return []


def filegrouping_columns_for_report_type(config: Dict[str, Any], report_type: Union[int, str]) -> List[str]:
    fields = config["fields"]

    def sc(logical_name: str) -> str:
        return fields[logical_name]["source_column"]

    _, type_key = parse_report_type_key(report_type)

    if type_key == "3_CONS":
        return [sc("consorzio")]
    if type_key == "3_GRUPPO":
        return [sc("gruppo_commerciale")]
    if type_key == "5":
        return [sc("agenzia_anagrafica")]

    return grouping_columns_for_report_type(config, report_type)


def preferred_totals_label_col(config: Dict[str, Any], df: pd.DataFrame) -> str:
    fields = config.get("fields", {})
    gm = fields.get("gruppo_merceologico", {}).get("source_column")
    if gm and gm in df.columns:
        return gm

    # Fallback to last non-numeric column
    non_num = [c for c in df.columns if c not in _CLIENT_NUMERIC_2DP_COLS]
    return non_num[-1] if non_num else df.columns[0]


def prepare_dataframe_for_report_type(
    config: Dict[str, Any],
    report_type: Union[int, str],
) -> Tuple[pd.DataFrame, str, str]:
    base_type, type_key = parse_report_type_key(report_type)

    logical_fields = get_logical_fields_specific_first(config, type_key)
    source_columns_map = get_source_columns_map(config, logical_fields)

    source_name = config["report_types"][type_key].get("source", "main_csv")
    df, input_yyyymmdd = load_source_dataframe_for_report_type(config, source_name, type_key)

    ordered_source_columns = [source_columns_map[lf] for lf in logical_fields]
    missing = [c for c in ordered_source_columns if c not in df.columns]

    if missing:
        raise RuntimeError(
            "Source CSV is missing columns: "
            + ", ".join(missing)
            + "\nCSV columns found: "
            + ", ".join(df.columns.astype(str))
        )

    df = df[ordered_source_columns].copy()
    ensure_unique_columns(df)

    df = coerce_numeric_2dp_columns(df)

    group_cols = aggregation_grouping_columns_for_report_type(config, type_key)
    df = aggregate_to_report_grain(df, group_cols=group_cols)

    roman = REPORT_TYPE_ROMAN[base_type]
    return df, input_yyyymmdd, roman


def build_report_excels_with_metadata(
    report_type: Union[int, str],
) -> List[Tuple[bytes, str, str, str]]:
    config: Dict[str, Any] = load_config()
    df, input_yyyymmdd, roman = prepare_dataframe_for_report_type(config, report_type)

    _, type_key = parse_report_type_key(report_type)

    subtotal_cols = grouping_columns_for_report_type(config, report_type)
    for c in subtotal_cols:
        if c not in df.columns:
            raise RuntimeError(f"Report type {report_type} requires grouping column '{c}' but it is missing")

    file_group_cols = filegrouping_columns_for_report_type(config, report_type)
    for c in file_group_cols:
        if c not in df.columns:
            raise RuntimeError(f"Report type {report_type} requires file grouping column '{c}' but it is missing")

    if not file_group_cols:
        label_col = preferred_totals_label_col(config, df)
        keep_cols: List[str] = []  # no group keys to keep
        df_out = append_totals_row(df.reset_index(drop=True), label_col=label_col, keep_cols=keep_cols)

        xlsx_bytes = to_excel_bytes(
            df_out,
            sheet_name=f"tipo_{type_key}",
            merge_columns=None,
        )
        return [(xlsx_bytes, input_yyyymmdd, roman, "")]

    if type_key == "1":
        grp_mer_col = config["fields"]["gruppo_merceologico"]["source_column"]
        if grp_mer_col not in df.columns:
            raise RuntimeError(f"Type 1 requires columns '{grp_mer_col}'")
        df = df.sort_values(by=[grp_mer_col], kind="mergesort")

    outputs: List[Tuple[bytes, str, str, str]] = []

    df_base = df.reset_index(drop=True)

    for keys, df_file in df_base.groupby(file_group_cols, dropna=False, sort=False):
        if not isinstance(keys, tuple):
            keys = (keys,)

        suffix_parts = [sanitize_for_filename(k) for k in keys]
        suffix = "__".join(suffix_parts)

        df_file_out = df_file.copy()

        label_col = preferred_totals_label_col(config, df_file_out)

        sort_cols = [c for c in subtotal_cols if c in df_file_out.columns]
        if sort_cols:
            df_file_out = df_file_out.sort_values(by=sort_cols, kind="mergesort").reset_index(drop=True)

        if len(subtotal_cols) >= 2 and len(file_group_cols) == 1:
            df_file_out = insert_subtotals(df_file_out, subtotal_cols=subtotal_cols, label_col=label_col)
        else:
            keep_cols = file_group_cols[:]  # keep group key(s)
            df_file_out = append_totals_row(df_file_out, label_col=label_col, keep_cols=keep_cols)

        xlsx_bytes = to_excel_bytes(
            df_file_out,
            sheet_name=f"tipo_{type_key}",
            merge_columns=file_group_cols,
        )

        outputs.append((xlsx_bytes, input_yyyymmdd, roman, suffix))

    if not outputs:
        raise RuntimeError(f"Type {report_type}: no output generated")

    return outputs
